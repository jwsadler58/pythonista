import json
from openpyxl import Workbook

#
# base classes and methods for pedal steel sliderule app
# Copyright (c) John W Sadler - October 2025
#
# Note names in a given key are such that accidentals in the key are either all sharps or all flats
# Brute force: predefine the note names of each key by degree
# Compute: for each key derive the note names 
# - break keys into sharp and flat ones, 
# - walk major scale intervals to get the note names
# map chromatic degree to note name in a key
# map chromatic offset
# map note name to offset
#
# Global State:
## Lefty or Righty
## Copedent store
#
# Copedent contains the configuration state of the instrument. 
# Initialized with any number of strings from 6 to 16 (limit imposed by UI design).
# Named copedents persist. Known copedents can be listed, edited, and stored
# State: 
## nStrings, 
## Key at nut for this tuning, 
## StringNotes, 
## PedalOffsets (dictionary of name : (string, offset))
#

_noteNames = {
	"C": 0, 
	"C#":1,  "Db":1, 
	"D" :2, 
	"D#":3,  "Eb":3, 
	"E" :4, 
	"F" :5, 
	"F#":6,  "Gb":6, 
	"G" :7, 
	"G#":8,  "Ab":8, 
	"A" :9, 
	"A#":10, "Bb":10, 
	"B" :11, "Cb":11
}


_majIndices  = [0,3,5,6,8,10,11]
_degreeNames = ["1", "b2", "2", "b3","3","4", "-5", "5","b6","6","b7","7"]	

def validKey(key):
	'''Return prettied key name, or raise ValueError. 
		Deals with caps, and checks valid key name. 
		Lowercase is not treated as minor'''
	key = key.capitalize()
	if not (key in _noteNames):
		raise ValueError("'"+key+"' is not in the set of valid keys") 
	return key

def degreeToNote(degree, key="C"):
	'''Returns note name of the given chromatic degree in the key (default key of C)'''
	_sharpKeys   = ["C","G","D" ,"A" ,"E" ,"B", "F#","C#"]
	sharpNotes  = ["C","C#","D","D#","E","F","F#","G","G#","A","A#","B"]
	flatNotes   = ["C","Db","D","Eb","E","F","Gb","G","Ab","A","Bb","B"]
	
	keyDegree = _noteNames[key]
	i = (keyDegree + degree) % 12
	if key in _sharpKeys:
		return sharpNotes[i]
	else:
		return flatNotes[i]

def noteToDegree(note, key="C"):
	''' returns the chromatic degree associated with the given note name and key '''
	note = note.capitalize()
	keyDegree = _noteNames[key]
	noteDegree = _noteNames[note]
	return (noteDegree - keyDegree) % 12


#	
# N O T E
#
class Note:
	'''Transposing note class
		- state includes chromatic degree, key, transpose offset
		- transposition changes the effective key, so note name changes but the degree does not
	'''
	
	def __init__(self, degree=0, key="C"):
		self.degree = degree
		self.key = key
		self.xpose = 0

	@classmethod
	def fromName(cls, name, key="C"):
		'''alternate constructor that takes the note name instead of degree'''
		degree = noteToDegree(name, key)
		return cls(degree, key)
		
	@classmethod
	def fromDegreeName(cls, degName, key="C"):
		degree = 	_degreeNames.index(degName)
		return cls(degree, key)
	
	def setTranspose(self, offset):
		self.xpose = offset

	def getName(self):
		'''returns the note name string of this offset, transpose, & key'''
		d = self.degree + self.xpose
		return degreeToNote(d, self.key)
		
	def getDegreeName(self):
		'''returns the name of the scale degree associated with this note (transpose ignored)'''
		return _degreeNames[self.degree]
		

#
# C O P E D E N T
#
# Key - nominal key of the copedent ("E" for E9th for example) - determines the degree names of notes
# nutNotes - names of notes at nut from nearest to farthest string - comma separated
# pedals - name:[(string#,shift),..] (name mapped to GUI later)
# -- name is unique to the pedal or lever - however you want to identify it
# -- mapped to a list of 1 or more tuples indicating string number and pedal offset in (integer) semitones
# -- string# is 0 for innermost string
#
class Copedent:
	def __init__(self, id="E9th", key="E", nutNotes="B,D,E,F#,G#,B,E,G#,D#,F#"):
		self.key = validKey(key)
		self.id = id
		self.nutNotes = nutNotes.split(',')
		self.pedals = {}
		self.reset()

	def reset(self):
		''' recalculate nutDegNames based on key and nutNotes '''
		self.nStrings = len(self.nutNotes)
		self.nutDegNames = []
		
		for noteName in self.nutNotes:
			n = Note.fromName(noteName, self.key)
			self.nutDegNames.append(n.getDegreeName())
		
	def addPedal(self, pedalName, tupleList):
		'''append a pedal definition to the list of pedals. no check for uniqueness'''
		self.pedals[pedalName]=tupleList
	
	def loadPedals(self, jsonStr=""):
		'''replace pedals dict with the provided JSON string'''
		
		e9th = '''{"A":  [[0, 2], [5, 2]], \
			"B":  [[4, 1], [7, 1]], \
			"C":  [[5, 2], [6, 2]], \
			"LL": [[2,-1],[6,-1]], \
			"LR": [[2,1], [6,1]], \
			"RL": [[1,-1],[8,-2]], \
			"RU": [[5,-1]], \
			"RR": [[3,1],[9,1]] \
			}'''
		
		if jsonStr == "":
			self.pedals = json.loads(e9th)
		else:
			self.pedals = json.loads(jsonStr)
	
	def dumps(self):
		'''Dump JSON describing the copedent'''
		d = {}
		d["id"]      = self.id
		d["key"]     = self.key
		d["nutNotes"]= self.nutNotes
		d["pedals"]  = self.pedals
		return json.dumps(d)
		
	def save(self):
		fname = self.id+".json"
		with open(fname, "w") as f:
			f.write(self.dumps())

	def load(self, fname):
		with open(fname, "r") as f:
			jsonString = f.read()
		d = json.loads(jsonString)
		self.id = d["id"]
		self.key = d["key"]
		self.nutNotes = d["nutNotes"]
		self.pedals = d["pedals"]
		self.nStrings = len(self.nutNotes)
		

	def print(self, refDeg=0):
		'''Prints a copedent in a conventional format
			- Referenced to the chromatic key degree provided (default 0, 0..11) 
		'''
		#empty array for degree values - nstrings+1 wide by nPedals+1 tall
		nPedals = len(self.pedals)
		cells = [["" for i in range(self.nStrings+1)] for j in range(nPedals+1)]
		
		# list degrees relative to the reference key provided
		refNote = Note(refDeg, self.key)
		refKey  = refNote.getName()
		print(f"\n{self.id:5} copedent referenced to {refKey}")
		
		# populate the nut row degrees based on refDeg
		cells[0][0] = "Nut"
		nutDegrees = []
		for n in range(len(self.nutNotes)):
			nn = Note.fromName(self.nutNotes[n], refKey)
			nutDegrees.append(nn.degree)
			cells[0][n+1] = nn.getDegreeName()
		
		# list degrees by pedal (relative to refDeg)		
		row = 1
		for ped,bendList in self.pedals.items():
			cells[row][0] = ped
			# assumes pedal offsets are ordered by acending string
			for snum,bend in bendList:
				offset = (nutDegrees[snum]+bend) % 12
				n = Note(offset, refKey)
				cells[row][snum+1] = n.getDegreeName()
			row += 1
		
		xpose = [list(row) for row in zip(*cells)]
		xpose.reverse()
		print("".join(f"{cell:5}" for cell in xpose.pop())) # header row
		for row in xpose:
			print("".join(f"{cell:5}" for cell in row))

	def toWorksheet(self, ws, refDeg=0):
		# Display copedent relative to the reference degree provided in the provided worksheet
		refNote = Note(refDeg, self.key)
		refKey  = refNote.getName()
		ws.title = f"\n{self.id} in {refKey}"
		ws.cell(row=2,column=2).value = f"\n{self.id} copedent relative to {refKey}"
		
		# populate the nut row degrees based on refDeg
		ws.cell(row=3,column=2).value = "Nut"
		nutDegrees = []
		for n in range(len(self.nutNotes)):
			nn = Note.fromName(self.nutNotes[n], refKey)
			nutDegrees.append(nn.degree)
			ws.cell(row=3,column=3+n).value = nn.getDegreeName()
		
		# list degrees by pedal (relative to refDeg)		
		row = 4
		for ped,bendList in self.pedals.items():
			ws.cell(row=row, column=2).value = ped
			# assumes pedal offsets are ordered by ascending string
			for snum,bend in bendList:
				offset = (nutDegrees[snum]+bend) % 12
				n = Note(offset, refKey)
				ws.cell(row=row, column=snum+3).value = n.getDegreeName()
			row += 1


#END CLASS COPEDENT

def detesto(fname):
	''' save detesto charts for the fname copedent to an xlsx file, one worksheet per chart'''
	#todo: Refactor or rewrite Print so that it can add cells to a worksheet
	c = Copedent()
	c.load(fname)	
	
	wb = Workbook()
	ws = wb.active
	
	noteSet = set(c.nutNotes)
	for noteName in noteSet:
		n = Note.fromName(noteName, c.key)
		c.print(n.degree)


if __name__ == '__main__':
	detesto("E9th.json")
	


